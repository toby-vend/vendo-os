"""Sword Stall P&L, Blue Sense 'four-block' layout, for an arbitrary run of months.

Same layout as build_pnl_sheet.py (Gross Revenue → Discount → Return → Shipping → Net Revenue → COGS →
Fulfilment & Shipping → Transaction Fees → Gross Margin → Direct Advertising → Contribution Margin → OpEx →
EBITDA → aMER → aROAS, with detail blocks and a %-of-revenue block) but months are a parameter and COGS is
taken per month from a ShopifyQL pull rather than pro-rated from a quarter total.

usage: python3 build_pnl_sheet_range.py <out.xlsx> <assumptions.json> --from=YYYY-MM --to=YYYY-MM [--no-raw]

Inputs
  data/sword-stall/uploads/shopify/ql/daily-sales-<from>-<to>.csv        ShopifyQL 'FROM sales … TIMESERIES day'
  data/sword-stall/uploads/shopify/ql/cogs-by-variant-by-month-*.csv     ShopifyQL sales by variant, one block per month
  vendo.db gads_campaign_spend (account 2310522325)                        Google Ads API
  data/sword-stall/uploads/meta/*.csv                                      Ads Manager campaign exports (monthly + daily)
  data/sword-stall/uploads/sword-stall.myshopify.com-all-*.csv             Triple Whale daily feed (cross-check only)
  vendo.db xero_invoices (contact 'The Sword Stall')                       Vendo fee, where synced
"""
import csv, re, sqlite3, json, sys, glob, calendar
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = "/Users/Toby_1/Vendo-OS"
QL = f"{ROOT}/data/sword-stall/uploads/shopify/ql"
TW_CSV = f"{ROOT}/data/sword-stall/uploads/sword-stall.myshopify.com-all-2026-01-01-2026-08-15.csv"
META_MONTHLY = f"{ROOT}/data/sword-stall/uploads/meta/The-Sword-Stall-Campaigns-Jul-16-2023-Aug-16-2026.csv"
META_DAILY = [f"{ROOT}/data/sword-stall/uploads/meta/The-Sword-Stall-Campaigns-Aug-1-2026-Aug-31-2026.csv"]

OUT = sys.argv[1]
cfg = json.load(open(sys.argv[2]))
arg = lambda k: next((a.split("=", 1)[1] for a in sys.argv if a.startswith(f"--{k}=")), None)
FROM, TO = arg("from"), arg("to")
NO_RAW = "--no-raw" in sys.argv
if not (FROM and TO): sys.exit("need --from=YYYY-MM --to=YYYY-MM")

def month_range(a, b):
    y, m = map(int, a.split("-")); out = []
    while True:
        out.append(f"{y:04d}-{m:02d}")
        if f"{y:04d}-{m:02d}" == b: return out
        m += 1
        if m == 13: y, m = y + 1, 1
MONTHS = month_range(FROM, TO); N = len(MONTHS)
MON = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
ML = {m: f"{MON[int(m[5:]) - 1]} {m[:4]}" for m in MONTHS}
PERIOD = f"{ML[MONTHS[0]]}–{ML[MONTHS[-1]]}" if MONTHS[0][:4] != MONTHS[-1][:4] else f"{MON[int(MONTHS[0][5:]) - 1]}–{MON[int(MONTHS[-1][5:]) - 1]} {MONTHS[0][:4]}"
PERIOD_LONG = PERIOD.replace("–", " – ")
first_day = f"{FROM}-01"; last_day = f"{TO}-{calendar.monthrange(int(TO[:4]), int(TO[5:]))[1]:02d}"
n = lambda x: float(x) if x not in ("", None) else 0.0
MC = [get_column_letter(2 + j) for j in range(N)]          # month columns
TOTC = get_column_letter(2 + N); BASC = get_column_letter(3 + N); SRCC = get_column_letter(4 + N)
BAS_I = 3 + N; SRC_I = 4 + N

# ---------------- data ----------------
DAILY_CSV = f"{QL}/daily-sales-{first_day}-{last_day}.csv"
COGS_CSV = f"{QL}/cogs-by-variant-by-month-{FROM}-{TO}.csv"
NAMES = {"orders": "Orders", "gross_sales": "Gross sales", "discounts": "Discounts", "sales_reversals": "Sales reversals",
         "net_sales": "Net sales", "shipping_charges": "Shipping charges", "duties": "Duties", "additional_fees": "Additional fees",
         "taxes": "Taxes", "total_sales": "Total sales"}
shop = defaultdict(lambda: defaultdict(float))
for row in csv.DictReader(open(DAILY_CSV)):
    m = row["day"][:7]
    if m in MONTHS:
        for k, v in NAMES.items(): shop[m][v] += n(row[k])
missing_months = [m for m in MONTHS if m not in shop]
if missing_months: sys.exit(f"daily CSV has no rows for {missing_months}")

con = sqlite3.connect(f"{ROOT}/data/vendo.db")
gads = {m: dict(spend=0, conv=0, value=0, synced="") for m in MONTHS}
for m, sp, cv, cval, syn in con.execute(
        "select substr(date,1,7), sum(spend), sum(conversions), sum(conversion_value), max(synced_at) "
        "from gads_campaign_spend where account_id like '%2310522325%' and date between ? and ? group by 1", (first_day, last_day)):
    gads[m] = dict(spend=sp, conv=cv, value=cval, synced=syn)
xero = defaultdict(dict)          # reference -> month -> (net, invoice numbers)
for ref_, m, sub, inv in con.execute(
        "select coalesce(reference,''), substr(date,1,7), sum(subtotal), group_concat(invoice_number) from xero_invoices "
        "where contact_name like '%Sword Stall%' and type='ACCREC' and date between ? and ? group by 1, 2", (first_day, last_day + "T23:59:59")):
    xero[ref_][m] = (sub, inv)
xero_last = con.execute("select max(substr(date,1,7)) from xero_invoices where contact_name like '%Sword Stall%' and type='ACCREC'").fetchone()[0]

money = lambda s: n((s or "").replace("£", "").replace(",", ""))
meta_am = defaultdict(float); meta_am_end = {}
for r in csv.DictReader(open(META_MONTHLY)):
    m = r["Reporting starts"][:7]; meta_am[m] += money(r["Amount spent (GBP)"]); meta_am_end[m] = max(meta_am_end.get(m, ""), r["Reporting ends"])
for f in META_DAILY:                       # a daily export for a month overrides the (possibly partial) monthly row
    daily = defaultdict(float)
    for r in csv.DictReader(open(f)): daily[r["Reporting starts"][:7]] += money(r["Amount spent (GBP)"])
    for m, v in daily.items(): meta_am[m] = v; meta_am_end[m] = "full month (daily export)"
meta_tw = defaultdict(float); tw_last = ""
for row in csv.DictReader(open(TW_CSV)):
    if row["Source"] == "Meta": meta_tw[row["Date"][:7]] += money(row["Ad Spend"]); tw_last = max(tw_last, row["Date"])
tw_full = {m for m in MONTHS if m in meta_tw and m < tw_last[:7]}   # months the TW feed covers end to end

rows = [r for r in csv.DictReader(open(COGS_CSV)) if r["month"] in MONTHS]
for r in rows:
    for k in ("product_title", "product_variant_title", "product_variant_sku"): r[k] = r[k] if r[k] not in ("", "None") else ""
def pref(s):
    m = re.match(r"[A-Za-z]+", s or ""); return m.group(0).upper() if m else "(no SKU)"
ref = defaultdict(lambda: [0.0, 0.0, set()])
for r in rows:
    if n(r["cost_of_goods_sold"]) > 0 and n(r["net_items_sold"]) > 0:
        p = pref(r["product_variant_sku"]); ref[p][0] += n(r["gross_sales"]); ref[p][1] += n(r["cost_of_goods_sold"]); ref[p][2].add(r["product_variant_sku"] or r["product_title"])
overall_ratio = sum(v[1] for v in ref.values()) / sum(v[0] for v in ref.values())
MIN_ROWS = 5
def estimate(r):
    title = r["product_title"].lower()
    if not r["product_title"]:
        return 0.0, "No product on the line (order-level refund / adjustment) – no cost", 0.0
    if "membership" in title or "gift card" in title or "subscription" in title:
        return 0.0, "Digital/membership – no cost of goods", 0.0
    if re.search(r"weta|hot toys|statue|bust|sculpture|pure ?arts|life[- ]size|prop replica|puppet|1:2 scale|1/6", title) \
            or pref(r["product_variant_sku"]) in ("HT", "WETA", "TTSP", "RLWB"):
        return n(r["gross_sales"]) / 1.10, "Collectible/statue – James: ~10% markup on statues (14 May 2026 call) → cost ≈ 91% of price", 1 / 1.10
    p = pref(r["product_variant_sku"])
    if len(ref[p][2]) >= MIN_ROWS:
        ratio = ref[p][1] / ref[p][0]
        return n(r["gross_sales"]) * ratio, f"Supplier-prefix '{p}' avg ({len(ref[p][2])} SKUs with cost)", ratio
    return n(r["gross_sales"]) * overall_ratio, "Store-wide avg COGS % of gross (no prefix reference)", overall_ratio
missing = [r for r in rows if n(r["cost_of_goods_sold"]) == 0 and n(r["net_items_sold"]) > 0 and r["product_title"]]
cogs_actual = {m: sum(n(r["cost_of_goods_sold"]) for r in rows if r["month"] == m) for m in MONTHS}
est_by_m = {m: sum(estimate(r)[0] for r in missing if r["month"] == m) for m in MONTHS}
net_by_m = {m: sum(n(r["net_sales"]) for r in rows if r["month"] == m) for m in MONTHS}
covered = {m: sum(n(r["net_sales_with_cost_recorded"]) for r in rows if r["month"] == m) / net_by_m[m] for m in MONTHS}
n_missing_variants = len({(r["product_title"], r["product_variant_title"], r["product_variant_sku"]) for r in missing})
csv_net = sum(net_by_m.values()); cogs_total = sum(cogs_actual.values()); est_total = sum(est_by_m.values())

# ---------------- styles ----------------
BOLD = Font(bold=True); H1 = Font(bold=True, size=14); WH = Font(bold=True, color="FFFFFF"); GREY = Font(color="6B7280", size=9)
F_HEAD = PatternFill("solid", fgColor="111827"); F_SECTION = PatternFill("solid", fgColor="E5E7EB"); F_TOTAL = PatternFill("solid", fgColor="D1FAE5")
F_ACT = PatternFill("solid", fgColor="FFFFFF"); F_EST = PatternFill("solid", fgColor="FEF3C7"); F_ASM = PatternFill("solid", fgColor="FDE68A"); F_UNK = PatternFill("solid", fgColor="FECACA")
F_INPUT = F_EST
BASIS_FILL = {"ACTUAL": F_ACT, "ESTIMATED": F_EST, "ASSUMED": F_ASM, "UNKNOWN": F_UNK, "DERIVED": F_ACT, "MIXED": F_EST}
GBP = '£#,##0;[Red]-£#,##0'; GBP2 = '£#,##0.00;[Red]-£#,##0.00'; PCT = '0.0%'
thin = Side(style="thin", color="9CA3AF"); TOP = Border(top=thin); DBL = Border(top=Side(style="double", color="111827"))
def hdr(ws, row, vals, fill=F_HEAD, font=WH, left=(1,)):
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v); c.fill = fill; c.font = font
        c.alignment = Alignment(horizontal="left" if i in left else "center", vertical="center", wrap_text=True)
def widths(ws, w):
    for i, x in enumerate(w, 1): ws.column_dimensions[get_column_letter(i)].width = x
def note(ws, row, text, span, h=30):
    c = ws.cell(row=row, column=1, value=text); c.alignment = Alignment(wrap_text=True, vertical="top"); c.font = GREY
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span); ws.row_dimensions[row].height = h
def monthly_values(it):
    """cfg 'monthly' may be a scalar, a list (one per month) or a {YYYY-MM: value} dict."""
    mv = it["monthly"]
    if isinstance(mv, dict): return [mv.get(m, mv.get("default", 0)) for m in MONTHS]
    if isinstance(mv, list): return mv
    return [mv] * N
def month_status(it):
    st = it["status"]
    return st if isinstance(st, list) else [st] * N
def basis_label(sts):
    return sts[0] if len(set(sts)) == 1 else "MIXED"

wb = Workbook()

# ================= Inputs =================
wsI = wb.active; wsI.title = "Inputs"
wsI["A1"] = "Inputs & assumptions – edit the yellow cells, the P&L recalculates"; wsI["A1"].font = H1
note(wsI, 2, "Status key: ACTUAL = hard source (Shopify, Google Ads API, Xero). ESTIMATED = figure James gave on a call, or derived from one. ASSUMED = industry default, no client data. UNKNOWN = never discussed, sitting at £0 so it neither flatters nor punishes the margin. MIXED = actual for some months, estimated for others (see note).", SRC_I, 34)
hdr(wsI, 4, ["Cost-of-delivery rates", "Value", "Unit", "Status"] + [""] * (N - 2) + ["Source / note"], left=(1, SRC_I))
INPUT = {}; r = 5
for it in cfg["inputs"]:
    wsI.cell(row=r, column=1, value=it["name"])
    c = wsI.cell(row=r, column=2, value=it["value"]); c.fill = F_INPUT
    c.number_format = PCT if it["unit"].startswith("%") else GBP2
    wsI.cell(row=r, column=3, value=it["unit"]); s = wsI.cell(row=r, column=4, value=it["status"]); s.fill = BASIS_FILL.get(it["status"], F_ACT)
    wsI.cell(row=r, column=SRC_I, value=it["note"]).alignment = Alignment(wrap_text=True, vertical="top")
    INPUT[it["key"]] = f"Inputs!$B${r}"; r += 1

# Xero-backed opex lines ("xero": <invoice reference>): actual where an invoice is synced for the month; for months
# after the last synced invoice the JSON fallback is carried forward (ESTIMATED); months before that with no invoice are £0 (ACTUAL).
for it in cfg["opex"]:
    if it.get("xero"):
        inv = xero.get(it["xero"], {}); vals, sts = [], []
        for m in MONTHS:
            if m in inv: vals.append(round(inv[m][0], 2)); sts.append("ACTUAL")
            elif m > xero_last: vals.append(it["monthly"]); sts.append("ESTIMATED")
            else: vals.append(0); sts.append("ACTUAL")
        it["monthly"] = vals; it["status"] = sts
        have = [m for m in MONTHS if m in inv]; lack = [m for m in MONTHS if m > xero_last]
        it["note"] = ((f"Xero: {', '.join(f'{inv[m][1]} {ML[m]} £{inv[m][0]:,.0f}' for m in have)} (net, + VAT), all paid. " if have else "")
                      + (f"{', '.join(ML[m] for m in lack)} not yet in the Xero sync (refresh token expired 4 Sep 2026 – run `npm run xero:auth`), carried at £{it['monthly'][-1]:,.0f}. " if lack else "")
                      + it["note"])

def block(title, items):
    global r
    r += 1; hdr(wsI, r, [title] + [ML[m] for m in MONTHS] + ["Status", "Source / note"], left=(1, SRC_I)); r += 1
    start = r
    for it in items:
        wsI.cell(row=r, column=1, value=it["name"])
        vals, sts = monthly_values(it), month_status(it)
        for j in range(N):
            c = wsI.cell(row=r, column=2 + j, value=vals[j]); c.fill = BASIS_FILL.get(sts[j], F_INPUT) if sts[j] == "ACTUAL" else F_INPUT; c.number_format = GBP
        s = wsI.cell(row=r, column=2 + N, value=basis_label(sts)); s.fill = BASIS_FILL.get(basis_label(sts), F_ACT)
        wsI.cell(row=r, column=3 + N, value=it["note"]).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    return start
DIRECT_START = block("Other direct marketing (monthly, ex VAT)", cfg["direct"])
OPEX_START = block("Operating expenses (monthly, ex VAT)", cfg["opex"])
widths(wsI, [46] + [13] * N + [12, 95])
wsI.column_dimensions[get_column_letter(3 + N)].width = 95   # note column of the monthly blocks

# ================= Revenue =================
wsR = wb.create_sheet("Revenue (Shopify)")
wsR["A1"] = f"Shopify sales by day (ShopifyQL, pulled 4 Sep 2026), summed by month – {PERIOD}"; wsR["A1"].font = H1
cols = ["Orders", "Gross sales", "Discounts", "Sales reversals", "Net sales", "Shipping charges", "Taxes", "Total sales"]
hdr(wsR, 3, ["Metric"] + [ML[m] for m in MONTHS] + [PERIOD])
REV = {}
for i, cn in enumerate(cols):
    rr = 4 + i; wsR.cell(row=rr, column=1, value=cn)
    for j, m in enumerate(MONTHS):
        c = wsR.cell(row=rr, column=2 + j, value=round(shop[m][cn], 2)); c.number_format = "#,##0" if cn == "Orders" else GBP
    c = wsR.cell(row=rr, column=2 + N, value=f"=SUM(B{rr}:{MC[-1]}{rr})"); c.number_format = "#,##0" if cn == "Orders" else GBP; c.font = BOLD
    REV[cn] = rr
note(wsR, 4 + len(cols) + 1, "Net sales = Gross − Discounts − Sales reversals (returns/refunds). Total sales = Net + Shipping + VAT. The P&L uses Net sales + Shipping charges, ex VAT. "
     f"Cross-check: the by-variant COGS pull sums to £{csv_net:,.0f} net sales for the same period (difference = gift cards / non-product items / report timing). "
     f"{ML[MONTHS[-1]]} was pulled on 4 Sep 2026 – late refunds will still move its 'Sales reversals' line.", 2 + N, 48)
widths(wsR, [40] + [15] * (N + 1))

# ================= Ad spend =================
wsA = wb.create_sheet("Ad Spend")
wsA["A1"] = f"In-platform spend – {PERIOD} (ex VAT)"; wsA["A1"].font = H1
hdr(wsA, 3, ["Platform / metric"] + [ML[m] for m in MONTHS] + [PERIOD, "Source"], left=(1, 3 + N))
def put(rr, label, vals, fmt, src="", bold=False):
    wsA.cell(row=rr, column=1, value=label).font = Font(bold=bold)
    for j, v in enumerate(vals):
        c = wsA.cell(row=rr, column=2 + j, value=v); c.number_format = fmt; c.font = Font(bold=bold)
    c = wsA.cell(row=rr, column=2 + N, value=f"=SUM(B{rr}:{MC[-1]}{rr})"); c.number_format = fmt; c.font = BOLD
    wsA.cell(row=rr, column=3 + N, value=src).alignment = Alignment(wrap_text=True, vertical="top")
synced = max((g["synced"] for g in gads.values()), default="")[:10]
put(4, "Google Ads spend", [round(gads[m]["spend"], 2) for m in MONTHS], GBP, f"Google Ads API → vendo.db (account 2310522325, synced {synced})", True)
put(5, "  Google conversions", [round(gads[m]["conv"]) for m in MONTHS], "#,##0", "Google Ads API")
put(6, "  Google conversion value (platform-attributed)", [round(gads[m]["value"], 2) for m in MONTHS], GBP, "Google Ads API – reference only")
put(7, "Meta Ads spend", [round(meta_am[m], 2) for m in MONTHS], GBP,
    "Ads Manager campaign exports (monthly rows to 16 Aug 2026, plus a daily export for the full month of Aug 2026). Meta API token expired 3 Sep 2026 – re-auth to pull live.", True)
put(8, "  Meta spend per Triple Whale daily feed (cross-check)", [round(meta_tw[m], 2) if m in tw_full else None for m in MONTHS], GBP,
    f"Triple Whale ad-spend feed to {tw_last}; matches Ads Manager to within £1 for every full month it covers. Blank where the feed does not cover the whole month.")
widths(wsA, [46] + [14] * (N + 1) + [90])

# ================= COGS estimated =================
wsE = wb.create_sheet("COGS – Estimated")
wsE["A1"] = "Variants sold with NO cost in Shopify – estimated cost, by month"; wsE["A1"].font = H1
note(wsE, 2, "Method: (1) memberships/gift cards = £0. (2) Statues, busts, Hot Toys, Weta, life-size props: James said statues carry ~10% markup vs ~70% on swords (14 May 2026), so cost = price ÷ 1.10. "
     f"(3) SKU prefix (supplier code) with ≥{MIN_ROWS} costed variants across {PERIOD}: that supplier's average COGS ÷ gross sales. (4) Otherwise the store-wide {overall_ratio:.1%}. Ratios in column I are editable; the P&L picks up column J by month.", 11, 46)
hdr(wsE, 4, ["Month", "Product title", "Variant", "SKU", "Units", "Gross sales", "Net sales", "Method", "Ratio", "Estimated COGS", "Confidence"], left=(2, 8))
for i, rw in enumerate(sorted(missing, key=lambda r: (r["month"], -n(r["gross_sales"])))):
    rr = 5 + i; est, method, ratio = estimate(rw)
    conf = "High" if ratio == 0 else ("Medium" if ("prefix" in method or "Collectible" in method) else "Low")
    vals = [rw["month"], rw["product_title"], rw["product_variant_title"], rw["product_variant_sku"], n(rw["net_items_sold"]), n(rw["gross_sales"]), n(rw["net_sales"]), method, round(ratio, 4), f"=F{rr}*I{rr}", conf]
    for j, v in enumerate(vals):
        c = wsE.cell(row=rr, column=1 + j, value=v)
        if j in (5, 6, 9): c.number_format = GBP2
        if j == 8: c.number_format = PCT; c.fill = F_INPUT
        if j == 9: c.fill = F_EST
lastE = 5 + len(missing) - 1; te = lastE + 2
wsE.cell(row=te, column=1, value="TOTAL").font = BOLD
for col in "EFGJ":
    c = wsE.cell(row=te, column=ord(col) - 64, value=f"=SUM({col}5:{col}{lastE})"); c.font = BOLD; c.border = TOP; c.number_format = GBP2 if col != "E" else "#,##0"
COGS_EST = lambda m: f"SUMIF('COGS – Estimated'!$A$5:$A${lastE},\"{m}\",'COGS – Estimated'!$J$5:$J${lastE})"
wsE.freeze_panes = "A5"; widths(wsE, [10, 58, 14, 16, 8, 13, 13, 60, 10, 15, 11])

# ================= COGS actual (raw) =================
fields = ["product_title", "product_variant_title", "product_variant_sku", "net_items_sold", "gross_sales", "discounts", "returns", "net_sales", "taxes", "total_sales", "cost_of_goods_sold", "net_sales_with_cost_recorded"]
FN = ["Product title", "Product variant title", "Product variant SKU", "Net items sold", "Gross sales", "Discounts", "Returns", "Net sales", "Taxes", "Total sales", "Cost of goods sold", "Net sales with cost recorded"]
if not NO_RAW:
    wsC = wb.create_sheet("COGS – Actual (raw)")
    wsC["A1"] = f"ShopifyQL 'sales by product variant' with cost of goods, one block per month, {PERIOD_LONG} (pulled 4 Sep 2026). Red rows = no cost recorded."; wsC["A1"].font = H1
    hdr(wsC, 3, ["Month"] + FN + ["Cost recorded?"], left=(2, 3, 4))
    for i, rw in enumerate(rows):
        rr = 4 + i; wsC.cell(row=rr, column=1, value=rw["month"])
        for j, f in enumerate(fields):
            v = rw[f] if f in ("product_title", "product_variant_title", "product_variant_sku") else n(rw[f])
            c = wsC.cell(row=rr, column=2 + j, value=v)
            if f not in ("product_title", "product_variant_title", "product_variant_sku", "net_items_sold"): c.number_format = GBP2
        flag = "NO" if (n(rw["cost_of_goods_sold"]) == 0 and n(rw["net_items_sold"]) > 0 and rw["product_title"]) else "yes"
        wsC.cell(row=rr, column=len(fields) + 2, value=flag)
        if flag == "NO":
            for j in range(len(fields) + 2): wsC.cell(row=rr, column=1 + j).fill = F_UNK
    last = 4 + len(rows) - 1; tr = last + 2
    wsC.cell(row=tr, column=1, value="TOTAL").font = BOLD
    for j, f in enumerate(fields):
        if f in ("product_title", "product_variant_title", "product_variant_sku"): continue
        col = get_column_letter(2 + j); c = wsC.cell(row=tr, column=2 + j, value=f"=SUM({col}4:{col}{last})"); c.font = BOLD; c.border = TOP; c.number_format = GBP2
    CG = get_column_letter(2 + fields.index("cost_of_goods_sold"))
    COGS_ACT = lambda m: f"SUMIF('COGS – Actual (raw)'!$A$4:$A${last},\"{m}\",'COGS – Actual (raw)'!${CG}$4:${CG}${last})"
    wsC.freeze_panes = "A4"; widths(wsC, [10, 60, 18, 18, 10, 13, 12, 12, 13, 12, 13, 14, 16, 12])
else:
    COGS_ACT = lambda m: str(round(cogs_actual[m], 2))

# ================= P&L =================
def build_pnl(name, use_est):
    ws = wb.create_sheet(name)
    ws["A1"] = f"THE SWORD STALL – P&L, {PERIOD_LONG}"; ws["A1"].font = H1
    ws["A2"] = ("COGS = Shopify recorded cost + estimate for the variants missing a cost" if use_est else "COGS = Shopify recorded cost only; variants with no cost sit at £0, so margin is overstated") + ". All figures ex VAT."
    ws["A2"].font = GREY
    legend = [("ACTUAL – Shopify / Google Ads API / Xero", F_ACT), ("ESTIMATED – from a client call or derived from one", F_EST), ("ASSUMED – industry default, no client data", F_ASM), ("UNKNOWN – never discussed, sitting at £0", F_UNK)]
    for i, (t, f) in enumerate(legend):
        c = ws.cell(row=3, column=1 + i, value=t); c.fill = f; c.font = Font(size=9, bold=True); c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    ws.row_dimensions[3].height = 30
    hdr(ws, 5, [""] + [ML[m] for m in MONTHS] + [PERIOD, "Basis", "Source / how it was calculated"], left=(1, SRC_I))
    R = "'Revenue (Shopify)'!"; A = "'Ad Spend'!"
    ALL = MC + [TOTC]
    rr = 6; ROW = {}
    def line(label, fN, basis="ACTUAL", src="", fmt=GBP, total=False, big=False, indent=False):
        nonlocal rr
        sts = basis if isinstance(basis, list) else [basis] * N
        lab = "" if total else basis_label(sts)
        c = ws.cell(row=rr, column=1, value=("    " if indent else "") + label); c.font = Font(bold=total, size=12 if big else 11)
        for j in range(N + 1):
            cell = ws.cell(row=rr, column=2 + j, value=fN[j]); cell.number_format = fmt; cell.font = Font(bold=total, size=12 if big else 11)
            if total: cell.fill = F_TOTAL; cell.border = TOP
            else: cell.fill = BASIS_FILL[sts[j] if j < N else lab]
        b = ws.cell(row=rr, column=BAS_I, value=lab); b.font = Font(size=9, bold=True); b.fill = F_TOTAL if total else BASIS_FILL[lab]
        if total: ws.cell(row=rr, column=1).fill = F_TOTAL; ws.cell(row=rr, column=1).border = TOP
        s = ws.cell(row=rr, column=SRC_I, value=src); s.font = GREY; s.alignment = Alignment(wrap_text=True, vertical="top")
        ROW[label] = rr; rr += 1; return rr - 1
    def section(t):
        nonlocal rr
        c = ws.cell(row=rr, column=1, value=t); c.font = BOLD
        for j in range(1, SRC_I + 1): ws.cell(row=rr, column=j).fill = F_SECTION
        rr += 1
    def gap():
        nonlocal rr; rr += 1
    rev_ref = lambda key: [f"={R}{c}{REV[key]}" for c in ALL]
    sumrow = lambda a, b: [f"=SUM({c}{a}:{c}{b})" for c in ALL]
    tot = lambda row: f"=SUM(B{row}:{MC[-1]}{row})"

    gr = line("Gross Revenue", rev_ref("Gross sales"), src="Shopify Analytics – gross sales before discounts and returns, ex VAT")
    di = line("Discount", rev_ref("Discounts"), src="Shopify Analytics – discounts (shown negative)")
    rt = line("Return", rev_ref("Sales reversals"), src="Shopify Analytics – sales reversals / refunds (shown negative)")
    sh = line("Shipping", rev_ref("Shipping charges"), src="Shopify Analytics – shipping charged to customers at checkout")
    rev = line("Net Revenue", sumrow(gr, sh), total=True, src="Gross − discounts − returns + shipping collected. The one revenue definition every KPI below uses.")
    cg = line("COGS", [f"={COGS_ACT(m)}" + (f"+{COGS_EST(m)}" if use_est else "") for m in MONTHS] + [tot(rr)],
              basis="ESTIMATED" if use_est else "ACTUAL",
              src=(f"Shopify recorded cost of goods per month (£{cogs_total:,.0f} for {PERIOD}, covering {min(covered.values()):.0%}–{max(covered.values()):.0%} of net sales each month)"
                   + (f" + £{est_total:,.0f} estimated for the {n_missing_variants} variants with no cost – see 'COGS – Estimated' tab" if use_est
                      else f". Variants with no cost recorded count as £0 here") + ". Real monthly COGS from ShopifyQL, no pro-rating."))
    fs = line("Fulfilment & Shipping", ["__FS__"] * (N + 1), basis="ESTIMATED", src="Carrier cost + packaging (detail below). Carrier: 14 May call, 'April alone, we did 6,000 just with DPD', £8–9/parcel on DPD retail, £6–7 expected on DX. Packaging never quantified.")
    tf = line("Transaction Fees", [f"={INPUT['txn_fee']}*{R}{c}{REV['Total sales']}" for c in ALL], basis="ASSUMED",
              src="Rate on Inputs × Shopify Total sales (amount actually charged incl. VAT & shipping). James: Patriot Payments took ~5%, most volume through PayPal. Needs the processor statements.")
    gm = line("Gross Margin", [f"={c}{rev}-{c}{cg}-{c}{fs}-{c}{tf}" for c in ALL], total=True, src="Net revenue − COGS − fulfilment & shipping − transaction fees (CM1)")
    dm = line("Direct Advertising & Marketing", ["__DM__"] * (N + 1), basis="ESTIMATED", src="Google + Meta in-platform spend (actual) + affiliate commission (estimated) – detail below. Agency fees and production sit in OpEx so only this line scales with revenue.")
    cm = line("Contribution Margin", [f"={c}{gm}-{c}{dm}" for c in ALL], total=True, src="Gross margin − direct advertising (CM2). Has to cover OpEx for the month to be profitable.")
    ox = line("OpEx", ["__OX__"] * (N + 1), basis="UNKNOWN", src="Only Vendo fee (actual to Jun) and the software James quoted (estimated) are in here. Rent, wages, packaging, Shopify/apps, accountancy are UNKNOWN and at £0 – detail below.")
    eb = line("EBITDA", [f"={c}{cm}-{c}{ox}" for c in ALL], total=True, big=True, src="Contribution margin − OpEx (net profit / CM3). Will fall as the UNKNOWN lines are filled in. James's accountant has Q2 at under 3%.")
    for j in range(1, SRC_I + 1): ws.cell(row=eb, column=j).border = DBL
    gap()
    am = line("aMER, %", [f"=IFERROR({c}{dm}/{c}{rev},0)" for c in ALL], fmt=PCT, src="Direct advertising ÷ net revenue (marketing efficiency on TOTAL revenue, not platform-attributed)")
    ar = line("aROAS, £", [f"=IFERROR({c}{rev}/{c}{dm},0)" for c in ALL], fmt='0.00', src="Net revenue ÷ direct advertising (blended, all channels, all revenue)")
    for x in (am, ar):
        ws.cell(row=x, column=BAS_I).value = ""
        for j in range(N + 2): ws.cell(row=x, column=2 + j).fill = F_ACT

    gap(); gap(); section("DETAIL – Fulfilment & Shipping")
    sf = line("Carrier cost (DPD → DX)", [f"={INPUT['ship_per_order']}*{R}{c}{REV['Orders']}" for c in ALL], basis="ESTIMATED", indent=True,
              src="Per-order rate on Inputs × orders. £6,000 DPD in April ÷ 950 orders = £6.32, plus Royal Mail for small items → £6.50. Needs the DPD / Royal Mail / DX invoices.")
    pk = line("Packaging & consumables", [f"={INPUT['pack_per_order']}*{R}{c}{REV['Orders']}" for c in ALL], basis="UNKNOWN", indent=True, src="Never quantified on a call. Boxes and foam for swords will not be zero.")
    fst = line("Fulfilment & Shipping total", [f"={c}{sf}+{c}{pk}" for c in ALL], total=True)
    for c in ALL: ws[f"{c}{fs}"].value = f"={c}{fst}"

    gap(); section("DETAIL – Direct Advertising & Marketing")
    g = line("Google Ads", [f"={A}{c}4" for c in ALL], indent=True, src="Google Ads API")
    mt = line("Meta Ads", [f"={A}{c}7" for c in ALL], indent=True, src="Ads Manager campaign export; Triple Whale feed agrees to within £1 for every full month it covers")
    for k, it in enumerate(cfg["direct"]):
        line(it["name"], [f"=Inputs!{MC[j]}{DIRECT_START + k}" for j in range(N)] + [tot(rr)], basis=month_status(it), indent=True, src=it["note"])
    dmt = line("Direct Advertising & Marketing total", [f"=SUM({c}{g}:{c}{rr - 1})" for c in ALL], total=True)
    for c in ALL: ws[f"{c}{dm}"].value = f"={c}{dmt}"

    gap(); section("DETAIL – OpEx")
    first_o = rr
    for k, it in enumerate(cfg["opex"]):
        line(it["name"], [f"=Inputs!{MC[j]}{OPEX_START + k}" for j in range(N)] + [tot(rr)], basis=month_status(it), indent=True, src=it["note"])
    oxt = line("OpEx total", [f"=SUM({c}{first_o}:{c}{rr - 1})" for c in ALL], total=True)
    for c in ALL: ws[f"{c}{ox}"].value = f"={c}{oxt}"

    gap(); gap()
    hdr(ws, rr, ["% OF NET REVENUE"] + [ML[m] for m in MONTHS] + [PERIOD, "", "The percentage view is what makes the P&L readable at a glance and comparable month to month"], left=(1, SRC_I)); rr += 1
    def pline(label, srcrow, bold=False, indent=False):
        nonlocal rr
        ws.cell(row=rr, column=1, value=("    " if indent else "") + label).font = Font(bold=bold)
        for c in ALL:
            x = ws[f"{c}{rr}"]; x.value = f"=IFERROR({c}{srcrow}/{c}${rev},0)"; x.number_format = PCT; x.font = Font(bold=bold)
            if bold: x.fill = F_TOTAL
        if bold: ws.cell(row=rr, column=1).fill = F_TOTAL
        rr += 1
    pline("Net Revenue", rev, True)
    pline("COGS", cg, indent=True); pline("Fulfilment & Shipping", fs, indent=True); pline("Transaction Fees", tf, indent=True)
    pline("Gross Margin", gm, True)
    pline("Google Ads", g, indent=True); pline("Meta Ads", mt, indent=True)
    for it in cfg["direct"]: pline(it["name"], ROW[it["name"]], indent=True)
    pline("Direct Advertising & Marketing", dm, True)
    pline("Contribution Margin", cm, True)
    pline("OpEx", ox, True)
    pline("EBITDA", eb, True)
    gap()
    ws.cell(row=rr, column=1, value="Marketing incl. Vendo fee (% of net revenue)").font = BOLD
    vrow = ROW[cfg["opex"][0]["name"]]
    for c in ALL:
        x = ws[f"{c}{rr}"]; x.value = f"=IFERROR(({c}{dm}+{c}{vrow})/{c}{rev},0)"; x.number_format = PCT; x.font = BOLD
    ws.cell(row=rr, column=SRC_I, value="What James's bookkeeper is likely calling 'marketing'. His P&L says 27%, with Vendo = 18% of revenue (17 Aug call).").font = GREY; rr += 1
    ws.cell(row=rr, column=1, value="Returns as % of gross revenue").font = BOLD
    for c in ALL:
        x = ws[f"{c}{rr}"]; x.value = f"=IFERROR(-{c}{rt}/{c}{gr},0)"; x.number_format = PCT; x.font = BOLD
    ws.cell(row=rr, column=SRC_I, value="Already deducted in net revenue. The biggest non-advertising profit lever.").font = GREY; rr += 1
    ws.cell(row=rr, column=1, value="Orders").font = BOLD
    for c in ALL:
        x = ws[f"{c}{rr}"]; x.value = f"={R}{c}{REV['Orders']}"; x.number_format = "#,##0"; x.font = BOLD
    rr += 1
    ws.cell(row=rr, column=1, value="EBITDA per order").font = BOLD
    for c in ALL:
        x = ws[f"{c}{rr}"]; x.value = f"=IFERROR({c}{eb}/{c}{rr - 1},0)"; x.number_format = GBP2; x.font = BOLD
    rr += 1
    ws.freeze_panes = "B6"; widths(ws, [48] + [13] * N + [15, 12, 95])
    return ws

build_pnl(f"P&L {PERIOD}", True)
build_pnl("P&L (recorded COGS only)", False)

# ================= Questions =================
wsQ = wb.create_sheet("Open questions for James")
wsQ["A1"] = "What is still needed to finish the P&L"; wsQ["A1"].font = H1
hdr(wsQ, 3, ["#", "Question", "Why it matters", "Placeholder in the model"], left=(2, 3, 4))
for i, q in enumerate(cfg["questions"]):
    for j, v in enumerate([i + 1] + q):
        c = wsQ.cell(row=4 + i, column=1 + j, value=v); c.alignment = Alignment(wrap_text=True, vertical="top")
widths(wsQ, [4, 72, 62, 40])

order = [f"P&L {PERIOD}", "P&L (recorded COGS only)", "Inputs", "Revenue (Shopify)", "Ad Spend", "COGS – Estimated"] + ([] if NO_RAW else ["COGS – Actual (raw)"]) + ["Open questions for James"]
wb._sheets = [wb[s] for s in order]
wb.save(OUT)
print("saved", OUT, "| months", N, "| recorded COGS", round(cogs_total), "| est gap", round(est_total), "| missing variant-months", len(missing), "| xero months", sorted(xero))
