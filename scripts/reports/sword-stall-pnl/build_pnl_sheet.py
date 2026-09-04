"""Sword Stall P&L in the Blue Sense 'four-block' layout: one line per item, months across,
pound block on top and a percentage-of-revenue block underneath, with a Basis column that says
whether each line is actual, estimated, assumed or unknown.

usage: python3 build_pnl_sheet.py <out.xlsx> <assumptions.json> [--no-raw]
"""
import csv, re, sqlite3, json, sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = "/Users/Toby_1/Vendo-OS"
COGS_CSV = "/Users/Toby_1/Downloads/Total sales by product variant w COGS - 2026-04-01 - 2026-06-30.csv"
SHOP_CSV = f"{ROOT}/data/sword-stall/uploads/shopify/Total sales over time - 2024-08-01 - 2026-08-16.csv"
TW_CSV = f"{ROOT}/data/sword-stall/uploads/sword-stall.myshopify.com-all-2026-01-01-2026-08-15.csv"
OUT = sys.argv[1]
cfg = json.load(open(sys.argv[2]))
NO_RAW = "--no-raw" in sys.argv
MONTHS = ["2026-04", "2026-05", "2026-06"]
ML = {"2026-04": "Apr 2026", "2026-05": "May 2026", "2026-06": "Jun 2026"}
n = lambda x: float(x) if x not in ("", None) else 0.0

# ---------------- data ----------------
shop = defaultdict(lambda: defaultdict(float))
with open(SHOP_CSV) as fh:
    r = csv.reader(fh); hdr = next(r)[:11]
    for row in r:
        m = row[0][:7]
        if m in MONTHS:
            for i, c in enumerate(hdr[1:], 1): shop[m][c] += n(row[i])

con = sqlite3.connect(f"{ROOT}/data/vendo.db")
gads = {}
for m, sp, cv, cval, syn in con.execute(
        "select substr(date,1,7), sum(spend), sum(conversions), sum(conversion_value), max(synced_at) "
        "from gads_campaign_spend where account_id like '%2310522325%' and date between '2026-04-01' and '2026-06-30' group by 1"):
    gads[m] = dict(spend=sp, conv=cv, value=cval, synced=syn)

meta_tw = defaultdict(float)
with open(TW_CSV) as fh:
    for row in csv.DictReader(fh):
        m = row["Date"][:7]
        if m in MONTHS and row["Source"] == "Meta":
            meta_tw[m] += n(row["Ad Spend"].replace("£", "").replace(",", ""))
meta_am = {"2026-04": 6119, "2026-05": 8277, "2026-06": 8314}

rows = list(csv.DictReader(open(COGS_CSV)))
def pref(s):
    m = re.match(r"[A-Za-z]+", s or ""); return m.group(0).upper() if m else "(no SKU)"
ref = defaultdict(lambda: [0.0, 0.0, 0])
for r in rows:
    if n(r["Cost of goods sold"]) > 0 and n(r["Net items sold"]) > 0:
        p = pref(r["Product variant SKU"]); ref[p][0] += n(r["Gross sales"]); ref[p][1] += n(r["Cost of goods sold"]); ref[p][2] += 1
overall_ratio = sum(v[1] for v in ref.values()) / sum(v[0] for v in ref.values())
MIN_ROWS = 5
def estimate(r):
    title = r["Product title"].lower()
    if "membership" in title or "gift card" in title or "subscription" in title:
        return 0.0, "Digital/membership – no cost of goods", 0.0
    if re.search(r"weta|hot toys|statue|bust|sculpture|pure ?arts|life[- ]size|prop replica|puppet|1:2 scale|1/6", title) \
            or pref(r["Product variant SKU"]) in ("HT", "WETA", "TTSP", "RLWB"):
        return n(r["Gross sales"]) / 1.10, "Collectible/statue – James: ~10% markup on statues (14 May 2026 call) → cost ≈ 91% of price", 1 / 1.10
    p = pref(r["Product variant SKU"])
    if ref[p][2] >= MIN_ROWS:
        ratio = ref[p][1] / ref[p][0]
        return n(r["Gross sales"]) * ratio, f"Supplier-prefix '{p}' avg ({ref[p][2]} SKUs with cost)", ratio
    return n(r["Gross sales"]) * overall_ratio, "Store-wide avg COGS % of gross (no prefix reference)", overall_ratio
missing = [r for r in rows if n(r["Cost of goods sold"]) == 0 and n(r["Net items sold"]) > 0]
cogs_actual = sum(n(r["Cost of goods sold"]) for r in rows)
net_missing = sum(n(r["Net sales"]) for r in missing)
items_missing = sum(n(r["Net items sold"]) for r in missing)
csv_net = sum(n(r["Net sales"]) for r in rows)
est_total = sum(estimate(r)[0] for r in missing)

# ---------------- styles ----------------
BOLD = Font(bold=True); H1 = Font(bold=True, size=14); WH = Font(bold=True, color="FFFFFF"); GREY = Font(color="6B7280", size=9)
F_HEAD = PatternFill("solid", fgColor="111827")
F_SECTION = PatternFill("solid", fgColor="E5E7EB")
F_TOTAL = PatternFill("solid", fgColor="D1FAE5")
F_ACT = PatternFill("solid", fgColor="FFFFFF")
F_EST = PatternFill("solid", fgColor="FEF3C7")   # amber = estimated from calls / derived
F_ASM = PatternFill("solid", fgColor="FDE68A")   # deeper amber = pure assumption
F_UNK = PatternFill("solid", fgColor="FECACA")   # red = unknown, at £0
F_INPUT = PatternFill("solid", fgColor="FEF3C7")
BASIS_FILL = {"ACTUAL": F_ACT, "ESTIMATED": F_EST, "ASSUMED": F_ASM, "UNKNOWN": F_UNK, "DERIVED": F_ACT}
GBP = '£#,##0;[Red]-£#,##0'; GBP2 = '£#,##0.00;[Red]-£#,##0.00'; PCT = '0.0%'
thin = Side(style="thin", color="9CA3AF"); TOP = Border(top=thin); DBL = Border(top=Side(style="double", color="111827"))
def hdr(ws, row, vals, fill=F_HEAD, font=WH):
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v); c.fill = fill; c.font = font
        c.alignment = Alignment(horizontal="left" if i in (1, 6) else "center", vertical="center", wrap_text=True)
def widths(ws, w):
    for i, x in enumerate(w, 1): ws.column_dimensions[get_column_letter(i)].width = x
def note(ws, row, text, span=6, h=30):
    c = ws.cell(row=row, column=1, value=text); c.alignment = Alignment(wrap_text=True, vertical="top"); c.font = GREY
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span); ws.row_dimensions[row].height = h

wb = Workbook()

# ================= Inputs =================
wsI = wb.active; wsI.title = "Inputs"
wsI["A1"] = "Inputs & assumptions – edit the yellow cells, the P&L recalculates"; wsI["A1"].font = H1
note(wsI, 2, "Status key: ACTUAL = hard source (Shopify, Google Ads API, Xero). ESTIMATED = figure James gave on a call, or derived from one. ASSUMED = industry default, no client data. UNKNOWN = never discussed, sitting at £0 so it neither flatters nor punishes the margin.", 6, 34)
hdr(wsI, 4, ["Cost-of-delivery rates", "Value", "Unit", "Status", "", "Source / note"])
INPUT = {}; r = 5
for it in cfg["inputs"]:
    wsI.cell(row=r, column=1, value=it["name"])
    c = wsI.cell(row=r, column=2, value=it["value"]); c.fill = F_INPUT
    c.number_format = PCT if it["unit"].startswith("%") else GBP2
    wsI.cell(row=r, column=3, value=it["unit"]); wsI.cell(row=r, column=4, value=it["status"])
    wsI.cell(row=r, column=6, value=it["note"]).alignment = Alignment(wrap_text=True, vertical="top")
    wsI.cell(row=r, column=4).fill = BASIS_FILL.get(it["status"], F_ACT)
    INPUT[it["key"]] = f"Inputs!$B${r}"; r += 1
def block(title, items):
    global r
    r += 1; hdr(wsI, r, [title, ML["2026-04"], ML["2026-05"], ML["2026-06"], "Status", "Source / note"]); r += 1
    start = r
    for it in items:
        wsI.cell(row=r, column=1, value=it["name"])
        for j in range(3):
            v = it["monthly"][j] if isinstance(it["monthly"], list) else it["monthly"]
            c = wsI.cell(row=r, column=2 + j, value=v); c.fill = F_INPUT; c.number_format = GBP
        s = wsI.cell(row=r, column=5, value=it["status"]); s.fill = BASIS_FILL.get(it["status"], F_ACT)
        wsI.cell(row=r, column=6, value=it["note"]).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    return start
DIRECT_START = block("Other direct marketing (monthly, ex VAT)", cfg["direct"])
OPEX_START = block("Operating expenses (monthly, ex VAT)", cfg["opex"])
widths(wsI, [46, 14, 14, 14, 12, 95])

# ================= Revenue =================
wsR = wb.create_sheet("Revenue (Shopify)")
wsR["A1"] = "Shopify Analytics – Total sales over time, summed by month"; wsR["A1"].font = H1
cols = ["Orders", "Gross sales", "Discounts", "Sales reversals", "Net sales", "Shipping charges", "Taxes", "Total sales"]
hdr(wsR, 3, ["Metric"] + [ML[m] for m in MONTHS] + ["Q2 2026"])
REV = {}
for i, cn in enumerate(cols):
    rr = 4 + i; wsR.cell(row=rr, column=1, value=cn)
    for j, m in enumerate(MONTHS):
        c = wsR.cell(row=rr, column=2 + j, value=round(shop[m][cn], 2)); c.number_format = "#,##0" if cn == "Orders" else GBP
    c = wsR.cell(row=rr, column=5, value=f"=SUM(B{rr}:D{rr})"); c.number_format = "#,##0" if cn == "Orders" else GBP; c.font = BOLD
    REV[cn] = rr
note(wsR, 4 + len(cols) + 1, "Net sales = Gross − Discounts − Sales reversals (returns/refunds). Total sales = Net + Shipping + VAT. The P&L uses Net sales + Shipping charges, ex VAT. "
     f"Cross-check: the product-variant COGS export sums to £{csv_net:,.0f} net sales for the same period (difference = gift cards / non-product items / report timing).", 5, 44)
widths(wsR, [40, 15, 15, 15, 15])

# ================= Ad spend =================
wsA = wb.create_sheet("Ad Spend")
wsA["A1"] = "In-platform spend – Apr–Jun 2026 (ex VAT)"; wsA["A1"].font = H1
hdr(wsA, 3, ["Platform / metric"] + [ML[m] for m in MONTHS] + ["Q2 2026", "Source"])
def put(rr, label, vals, fmt, src="", bold=False):
    wsA.cell(row=rr, column=1, value=label).font = Font(bold=bold)
    for j, v in enumerate(vals):
        c = wsA.cell(row=rr, column=2 + j, value=v); c.number_format = fmt; c.font = Font(bold=bold)
    c = wsA.cell(row=rr, column=5, value=f"=SUM(B{rr}:D{rr})"); c.number_format = fmt; c.font = BOLD
    wsA.cell(row=rr, column=6, value=src).alignment = Alignment(wrap_text=True, vertical="top")
put(4, "Google Ads spend", [round(gads[m]["spend"], 2) for m in MONTHS], GBP,
    f"Google Ads API → vendo.db (account 2310522325, synced {max(g['synced'] for g in gads.values())[:10]})", True)
put(5, "  Google conversions", [round(gads[m]["conv"]) for m in MONTHS], "#,##0", "Google Ads API")
put(6, "  Google conversion value (platform-attributed)", [round(gads[m]["value"], 2) for m in MONTHS], GBP, "Google Ads API – reference only")
put(7, "Meta Ads spend", [round(meta_tw[m], 2) for m in MONTHS], GBP,
    "Triple Whale daily ad-spend feed (sword-stall.myshopify.com-all-2026-01-01-2026-08-15.csv), cross-checked with the Ads Manager export below. Meta API token expired 3 Sep 2026 – re-auth to pull live.", True)
put(8, "  Meta spend per Ads Manager campaign export", [meta_am[m] for m in MONTHS], GBP, "Ads Manager export to 16 Aug 2026")
widths(wsA, [46, 14, 14, 14, 14, 90])

# ================= COGS estimated =================
wsE = wb.create_sheet("COGS – Estimated")
wsE["A1"] = "Variants sold with NO cost in Shopify – estimated cost"; wsE["A1"].font = H1
note(wsE, 2, "Method: (1) memberships/gift cards = £0. (2) Statues, busts, Hot Toys, Weta, life-size props: James said statues carry ~10% markup vs ~70% on swords (14 May 2026), so cost = price ÷ 1.10. "
     f"(3) SKU prefix (supplier code) with ≥5 costed variants: that supplier's average COGS ÷ gross sales. (4) Otherwise the store-wide {overall_ratio:.1%}. Ratios in column H are editable.", 10, 46)
hdr(wsE, 4, ["Product title", "Variant", "SKU", "Units", "Gross sales", "Net sales", "Method", "Ratio", "Estimated COGS", "Confidence"])
for i, rw in enumerate(sorted(missing, key=lambda r: -n(r["Gross sales"]))):
    rr = 5 + i; est, method, ratio = estimate(rw)
    conf = "High" if ratio == 0 else ("Medium" if ("prefix" in method or "Collectible" in method) else "Low")
    vals = [rw["Product title"], rw["Product variant title"], rw["Product variant SKU"], n(rw["Net items sold"]), n(rw["Gross sales"]), n(rw["Net sales"]), method, round(ratio, 4), f"=E{rr}*H{rr}", conf]
    for j, v in enumerate(vals):
        c = wsE.cell(row=rr, column=1 + j, value=v)
        if j in (4, 5, 8): c.number_format = GBP2
        if j == 7: c.number_format = PCT; c.fill = F_INPUT
        if j == 8: c.fill = F_EST
lastE = 5 + len(missing) - 1; te = lastE + 2
wsE.cell(row=te, column=1, value="TOTAL estimated COGS for the missing variants").font = BOLD
for col in "DEFI":
    c = wsE.cell(row=te, column=ord(col) - 64, value=f"=SUM({col}5:{col}{lastE})"); c.font = BOLD; c.border = TOP; c.number_format = GBP2 if col != "D" else "#,##0"
COGS_EST = f"'COGS – Estimated'!$I${te}"
wsE.freeze_panes = "A5"; widths(wsE, [58, 14, 16, 8, 13, 13, 60, 10, 15, 11])

# ================= COGS actual (raw) =================
if not NO_RAW:
    wsC = wb.create_sheet("COGS – Actual (raw)")
    wsC["A1"] = "Shopify 'Total sales by product variant with COGS', 1 Apr – 30 Jun 2026, as exported. Red rows = no cost recorded."; wsC["A1"].font = H1
    fields = list(rows[0].keys()); hdr(wsC, 3, fields + ["Cost recorded?"])
    for i, rw in enumerate(rows):
        rr = 4 + i
        for j, f in enumerate(fields):
            v = rw[f] if f in ("Product title", "Product variant title", "Product variant SKU") else n(rw[f])
            c = wsC.cell(row=rr, column=1 + j, value=v)
            if f not in ("Product title", "Product variant title", "Product variant SKU", "Net items sold"): c.number_format = GBP2
        flag = "NO" if (n(rw["Cost of goods sold"]) == 0 and n(rw["Net items sold"]) > 0) else "yes"
        wsC.cell(row=rr, column=len(fields) + 1, value=flag)
        if flag == "NO":
            for j in range(len(fields) + 1): wsC.cell(row=rr, column=1 + j).fill = F_UNK
    last = 4 + len(rows) - 1; tr = last + 2
    wsC.cell(row=tr, column=1, value="TOTAL").font = BOLD
    for j, f in enumerate(fields):
        if f in ("Product title", "Product variant title", "Product variant SKU"): continue
        col = get_column_letter(1 + j); c = wsC.cell(row=tr, column=1 + j, value=f"=SUM({col}4:{col}{last})"); c.font = BOLD; c.border = TOP; c.number_format = GBP2
    COGS_ACT = f"'COGS – Actual (raw)'!{get_column_letter(1 + fields.index('Cost of goods sold'))}{tr}"
    wsC.freeze_panes = "A4"; widths(wsC, [60, 18, 18, 10, 13, 12, 12, 13, 12, 13, 14, 16, 12])
else:
    COGS_ACT = str(round(cogs_actual, 2))

# ================= P&L (video layout) =================
def build_pnl(name, use_est):
    ws = wb.create_sheet(name)
    ws["A1"] = "THE SWORD STALL – P&L, Q2 2026 (April – June)"; ws["A1"].font = H1
    ws["A2"] = ("COGS = Shopify recorded cost + estimate for the 63 variants missing a cost" if use_est else "COGS = Shopify recorded cost only; the 63 variants with no cost sit at £0, so margin is overstated") + ". All figures ex VAT."
    ws["A2"].font = GREY
    legend = [("ACTUAL – Shopify / Google Ads API / Xero", F_ACT), ("ESTIMATED – from a client call or derived from one", F_EST), ("ASSUMED – industry default, no client data", F_ASM), ("UNKNOWN – never discussed, sitting at £0", F_UNK)]
    for i, (t, f) in enumerate(legend):
        c = ws.cell(row=3, column=1 + i, value=t); c.fill = f; c.font = Font(size=9, bold=True); c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    ws.row_dimensions[3].height = 30
    hdr(ws, 5, ["", ML["2026-04"], ML["2026-05"], ML["2026-06"], "Q2 2026", "Basis", "Source / how it was calculated"])
    R = "'Revenue (Shopify)'!"; A = "'Ad Spend'!"
    mc = ["B", "C", "D", "E"]
    rr = 6; ROW = {}
    def line(label, f4, basis="ACTUAL", src="", fmt=GBP, total=False, big=False, indent=False):
        nonlocal rr
        c = ws.cell(row=rr, column=1, value=("    " if indent else "") + label); c.font = Font(bold=total, size=12 if big else 11)
        for j in range(4):
            cell = ws.cell(row=rr, column=2 + j, value=f4[j]); cell.number_format = fmt; cell.font = Font(bold=total, size=12 if big else 11)
            if total: cell.fill = F_TOTAL; cell.border = TOP
            else: cell.fill = BASIS_FILL[basis]
        b = ws.cell(row=rr, column=6, value="" if total else basis); b.font = Font(size=9, bold=True)
        b.fill = F_TOTAL if total else BASIS_FILL[basis]
        if total: ws.cell(row=rr, column=1).fill = F_TOTAL; ws.cell(row=rr, column=1).border = TOP
        s = ws.cell(row=rr, column=7, value=src); s.font = GREY; s.alignment = Alignment(wrap_text=True, vertical="top")
        ROW[label] = rr; rr += 1; return rr - 1
    def section(t):
        nonlocal rr
        c = ws.cell(row=rr, column=1, value=t); c.font = BOLD
        for j in range(1, 8): ws.cell(row=rr, column=j).fill = F_SECTION
        rr += 1
    def gap():
        nonlocal rr; rr += 1

    # ---- detail rows are referenced by the summary, so reserve their positions first ----
    # Summary block (matches the Blue Sense sheet row for row)
    gr = line("Gross Revenue", [f"={R}{mc[j]}{REV['Gross sales']}" for j in range(4)], src="Shopify Analytics – gross sales before discounts and returns, ex VAT")
    di = line("Discount", [f"={R}{mc[j]}{REV['Discounts']}" for j in range(4)], src="Shopify Analytics – discounts (shown negative)")
    rt = line("Return", [f"={R}{mc[j]}{REV['Sales reversals']}" for j in range(4)], src="Shopify Analytics – sales reversals / refunds (shown negative)")
    sh = line("Shipping", [f"={R}{mc[j]}{REV['Shipping charges']}" for j in range(4)], src="Shopify Analytics – shipping charged to customers at checkout")
    rev = line("Net Revenue", [f"=SUM({col}{gr}:{col}{sh})" for col in "BCDE"], total=True, src="Gross − discounts − returns + shipping collected. The one revenue definition every KPI below uses.")
    q2c = f"{COGS_ACT}" + (f"+{COGS_EST}" if use_est else "")
    cg = line("COGS", [f"=({q2c})*{col}{ROW['Net Revenue']}/$E${ROW['Net Revenue']}" for col in "BCD"] + [f"={q2c}"],
              basis="ESTIMATED" if use_est else "ACTUAL",
              src=(f"Shopify recorded COGS £{cogs_actual:,.0f} (covers 96% of net sales)" + (f" + £{est_total:,.0f} estimated for {len(missing)} variants with no cost – see 'COGS – Estimated' tab" if use_est else f". {len(missing)} variants / £{net_missing:,.0f} net sales have no cost recorded and count as £0 here") + ". Shopify exports COGS as a quarter total, so months are allocated pro-rata to net revenue."))
    # Fulfilment = carrier + packaging (detail below)
    fs = line("Fulfilment & Shipping", ["__FS__"] * 4, basis="ESTIMATED", src="Carrier cost + packaging (detail below). Carrier: 14 May call, 'April alone, we did 6,000 just with DPD', £8–9/parcel on DPD retail, £6–7 expected on DX. Packaging never quantified.")
    tf = line("Transaction Fees", [f"={INPUT['txn_fee']}*{R}{mc[j]}{REV['Total sales']}" for j in range(4)], basis="ASSUMED",
              src="Rate on Inputs × Shopify Total sales (amount actually charged incl. VAT & shipping). James: Patriot Payments took ~5%, most volume through PayPal. Needs the Q2 processor statements.")
    gm = line("Gross Margin", [f"={col}{rev}-{col}{cg}-{col}{fs}-{col}{tf}" for col in "BCDE"], total=True, src="Net revenue − COGS − fulfilment & shipping − transaction fees (CM1)")
    dm = line("Direct Advertising & Marketing", ["__DM__"] * 4, basis="ESTIMATED", src="Google + Meta in-platform spend (actual) + affiliate commission (estimated) – detail below. Agency fees and production sit in OpEx so only this line scales with revenue.")
    cm = line("Contribution Margin", [f"={col}{gm}-{col}{dm}" for col in "BCDE"], total=True, src="Gross margin − direct advertising (CM2). Has to cover OpEx for the month to be profitable.")
    ox = line("OpEx", ["__OX__"] * 4, basis="UNKNOWN", src="Only Vendo fee (actual) and the software James quoted (estimated) are in here. Rent, wages, packaging, Shopify/apps, accountancy are UNKNOWN and at £0 – detail below.")
    eb = line("EBITDA", [f"={col}{cm}-{col}{ox}" for col in "BCDE"], total=True, big=True, src="Contribution margin − OpEx (net profit / CM3). Will fall as the UNKNOWN lines are filled in. James's accountant has Q2 at under 3%.")
    for col in "ABCDEFG": ws[f"{col}{eb}"].border = DBL
    gap()
    am = line("aMER, %", [f"=IFERROR({col}{dm}/{col}{rev},0)" for col in "BCDE"], fmt=PCT, src="Direct advertising ÷ net revenue (marketing efficiency on TOTAL revenue, not platform-attributed)")
    ws[f"F{am}"].value = ""; ws[f"F{am}"].fill = F_ACT
    for col in "BCDE": ws[f"{col}{am}"].fill = F_ACT
    ar = line("aROAS, £", [f"=IFERROR({col}{rev}/{col}{dm},0)" for col in "BCDE"], fmt='0.00', src="Net revenue ÷ direct advertising (blended, all channels, all revenue)")
    ws[f"F{ar}"].value = ""; ws[f"F{ar}"].fill = F_ACT
    for col in "BCDE": ws[f"{col}{ar}"].fill = F_ACT

    # ---- detail blocks ----
    gap(); gap(); section("DETAIL – Fulfilment & Shipping")
    sf = line("Carrier cost (DPD → DX)", [f"={INPUT['ship_per_order']}*{R}{mc[j]}{REV['Orders']}" for j in range(4)], basis="ESTIMATED", indent=True,
              src="Per-order rate on Inputs × orders. £6,000 DPD in April ÷ 950 orders = £6.32, plus Royal Mail for small items → £6.50. Needs the DPD / Royal Mail / DX invoices.")
    pk = line("Packaging & consumables", [f"={INPUT['pack_per_order']}*{R}{mc[j]}{REV['Orders']}" for j in range(4)], basis="UNKNOWN", indent=True, src="Never quantified on a call. Boxes and foam for swords will not be zero.")
    fst = line("Fulfilment & Shipping total", [f"={col}{sf}+{col}{pk}" for col in "BCDE"], total=True)
    for col in "BCDE": ws[f"{col}{fs}"].value = f"={col}{fst}"

    gap(); section("DETAIL – Direct Advertising & Marketing")
    g = line("Google Ads", [f"={A}{mc[j]}4" for j in range(4)], indent=True, src="Google Ads API")
    mt = line("Meta Ads", [f"={A}{mc[j]}7" for j in range(4)], indent=True, src="Triple Whale daily feed, matches the Ads Manager export to the pound")
    for k, it in enumerate(cfg["direct"]):
        line(it["name"], [f"=Inputs!{'BCD'[j]}{DIRECT_START + k}" for j in range(3)] + [f"=SUM(B{rr}:D{rr})"], basis=it["status"], indent=True, src=it["note"])
    dmt = line("Direct Advertising & Marketing total", [f"=SUM({col}{g}:{col}{rr - 1})" for col in "BCDE"], total=True)
    for col in "BCDE": ws[f"{col}{dm}"].value = f"={col}{dmt}"

    gap(); section("DETAIL – OpEx")
    first_o = rr
    for k, it in enumerate(cfg["opex"]):
        line(it["name"], [f"=Inputs!{'BCD'[j]}{OPEX_START + k}" for j in range(3)] + [f"=SUM(B{rr}:D{rr})"], basis=it["status"], indent=True, src=it["note"])
    oxt = line("OpEx total", [f"=SUM({col}{first_o}:{col}{rr - 1})" for col in "BCDE"], total=True)
    for col in "BCDE": ws[f"{col}{ox}"].value = f"={col}{oxt}"

    # ---- percentage breakdown along the bottom ----
    gap(); gap()
    hdr(ws, rr, ["% OF NET REVENUE", ML["2026-04"], ML["2026-05"], ML["2026-06"], "Q2 2026", "", "The percentage view is what makes the P&L readable at a glance and comparable month to month"]); rr += 1
    def pline(label, srcrow, bold=False, indent=False):
        nonlocal rr
        ws.cell(row=rr, column=1, value=("    " if indent else "") + label).font = Font(bold=bold)
        for col in "BCDE":
            c = ws[f"{col}{rr}"]; c.value = f"=IFERROR({col}{srcrow}/{col}${rev},0)"; c.number_format = PCT; c.font = Font(bold=bold)
            if bold: c.fill = F_TOTAL
        if bold: ws.cell(row=rr, column=1).fill = F_TOTAL
        rr += 1
    pline("Net Revenue", rev, True)
    pline("COGS", cg, indent=True); pline("Fulfilment & Shipping", fs, indent=True); pline("Transaction Fees", tf, indent=True)
    pline("Gross Margin", gm, True)
    pline("Google Ads", g, indent=True); pline("Meta Ads", mt, indent=True)
    for it in cfg["direct"]: pline(it["name"], ROW[it["name"]], indent=True)
    pline("Direct Advertising & Marketing", dm, True)
    pline("Contribution Margin", cm, True)
    pline("OpEx", ox, True)
    pline("EBITDA", eb, True)
    gap()
    ws.cell(row=rr, column=1, value="Marketing incl. Vendo fee (% of net revenue)").font = BOLD
    vrow = ROW[cfg["opex"][0]["name"]]
    for col in "BCDE":
        c = ws[f"{col}{rr}"]; c.value = f"=IFERROR(({col}{dm}+{col}{vrow})/{col}{rev},0)"; c.number_format = PCT; c.font = BOLD
    ws.cell(row=rr, column=7, value="What James's bookkeeper is likely calling 'marketing'. His P&L says 27%, with Vendo = 18% of revenue (17 Aug call).").font = GREY; rr += 1
    ws.cell(row=rr, column=1, value="Returns as % of gross revenue").font = BOLD
    for col in "BCDE":
        c = ws[f"{col}{rr}"]; c.value = f"=IFERROR(-{col}{rt}/{col}{gr},0)"; c.number_format = PCT; c.font = BOLD
    ws.cell(row=rr, column=7, value="Already deducted in net revenue. The biggest non-advertising profit lever.").font = GREY; rr += 1
    ws.cell(row=rr, column=1, value="Orders").font = BOLD
    for j, col in enumerate("BCDE"):
        c = ws[f"{col}{rr}"]; c.value = f"={R}{mc[j]}{REV['Orders']}"; c.number_format = "#,##0"; c.font = BOLD
    rr += 1
    ws.cell(row=rr, column=1, value="EBITDA per order").font = BOLD
    for col in "BCDE":
        c = ws[f"{col}{rr}"]; c.value = f"=IFERROR({col}{eb}/{col}{rr - 1},0)"; c.number_format = GBP2; c.font = BOLD
    rr += 1
    ws.freeze_panes = "B6"; widths(ws, [48, 14, 14, 14, 15, 12, 95])
    return ws

build_pnl("P&L Q2 2026", True)
build_pnl("P&L (recorded COGS only)", False)

# ================= Questions =================
wsQ = wb.create_sheet("Open questions for James")
wsQ["A1"] = "What is still needed to finish the P&L"; wsQ["A1"].font = H1
hdr(wsQ, 3, ["#", "Question", "Why it matters", "Placeholder in the model"])
for i, q in enumerate(cfg["questions"]):
    for j, v in enumerate([i + 1] + q):
        c = wsQ.cell(row=4 + i, column=1 + j, value=v); c.alignment = Alignment(wrap_text=True, vertical="top")
widths(wsQ, [4, 72, 62, 40])

order = ["P&L Q2 2026", "P&L (recorded COGS only)", "Inputs", "Revenue (Shopify)", "Ad Spend", "COGS – Estimated"] + ([] if NO_RAW else ["COGS – Actual (raw)"]) + ["Open questions for James"]
wb._sheets = [wb[s] for s in order]
wb.save(OUT)
print("saved", OUT, "| recorded COGS", round(cogs_actual), "| est gap", round(est_total), "| missing variants", len(missing))
