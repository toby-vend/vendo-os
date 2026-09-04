import csv, re, sqlite3, json, sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT="/Users/Toby_1/Vendo-OS"
COGS_CSV="/Users/Toby_1/Downloads/Total sales by product variant w COGS - 2026-04-01 - 2026-06-30.csv"
SHOP_CSV=f"{ROOT}/data/sword-stall/uploads/shopify/Total sales over time - 2024-08-01 - 2026-08-16.csv"
TW_CSV=f"{ROOT}/data/sword-stall/uploads/sword-stall.myshopify.com-all-2026-01-01-2026-08-15.csv"
OUT=sys.argv[1]
MONTHS=["2026-04","2026-05","2026-06"]
MLABEL={"2026-04":"Apr 2026","2026-05":"May 2026","2026-06":"Jun 2026"}
cfg=json.load(open(sys.argv[2]))   # assumptions + opex

n=lambda x: float(x) if x not in ('',None) else 0.0

# ---------- Shopify monthly ----------
shop=defaultdict(lambda: defaultdict(float))
with open(SHOP_CSV) as fh:
    r=csv.reader(fh); hdr=next(r)[:11]
    for row in r:
        m=row[0][:7]
        if m in MONTHS:
            for i,c in enumerate(hdr[1:],1): shop[m][c]+=n(row[i])

# ---------- Google Ads (DB) ----------
con=sqlite3.connect(f"{ROOT}/data/vendo.db")
gads={}
for m,sp,cv,cval,syn in con.execute("select substr(date,1,7), sum(spend), sum(conversions), sum(conversion_value), max(synced_at) from gads_campaign_spend where account_id like '%2310522325%' and date between '2026-04-01' and '2026-06-30' group by 1"):
    gads[m]=dict(spend=sp,conv=cv,value=cval,synced=syn)
# ---------- Meta (TW daily feed, cross-checked with Ads Manager export) ----------
meta_tw=defaultdict(float)
with open(TW_CSV) as fh:
    for row in csv.DictReader(fh):
        m=row['Date'][:7]
        if m in MONTHS and row['Source']=='Meta':
            meta_tw[m]+=n(row['Ad Spend'].replace('£','').replace(',',''))
meta_am={"2026-04":6119,"2026-05":8277,"2026-06":8314}  # Ads Manager campaign export (to 16 Aug 2026), from outputs/analyses/2026-08-16-sword-stall-yoy.md

# ---------- COGS CSV ----------
rows=list(csv.DictReader(open(COGS_CSV)))
def pref(s):
    m=re.match(r'[A-Za-z]+',s or ''); return m.group(0).upper() if m else '(no SKU)'
ref=defaultdict(lambda:[0.0,0.0,0])
for r in rows:
    if n(r["Cost of goods sold"])>0 and n(r["Net items sold"])>0:
        p=pref(r["Product variant SKU"]); ref[p][0]+=n(r["Gross sales"]); ref[p][1]+=n(r["Cost of goods sold"]); ref[p][2]+=1
tot_gross_wc=sum(v[0] for v in ref.values()); tot_cogs_wc=sum(v[1] for v in ref.values())
overall_ratio=tot_cogs_wc/tot_gross_wc
MIN_ROWS=5
def estimate(r):
    """returns (est_cogs, method, ratio)"""
    title=r["Product title"].lower()
    if "membership" in title or "gift card" in title or "subscription" in title:
        return 0.0,"Digital/membership – no cost of goods",0.0
    if re.search(r"weta|hot toys|statue|bust|sculpture|pure ?arts|life[- ]size|prop replica|puppet|1:2 scale|1/6", title) or pref(r["Product variant SKU"]) in ("HT","WETA","TTSP","RLWB"):
        return n(r["Gross sales"])/1.10, "Collectible/statue – James: ~10% markup on statues (14 May 2026 call) → cost ≈ 91% of price", 1/1.10
    p=pref(r["Product variant SKU"])
    if ref[p][2]>=MIN_ROWS:
        ratio=ref[p][1]/ref[p][0]; return n(r["Gross sales"])*ratio, f"Supplier-prefix '{p}' avg ({ref[p][2]} SKUs with cost)", ratio
    return n(r["Gross sales"])*overall_ratio, "Store-wide avg COGS % of gross (no prefix reference)", overall_ratio

missing=[r for r in rows if n(r["Cost of goods sold"])==0 and n(r["Net items sold"])>0]
est_total=sum(estimate(r)[0] for r in missing)
cogs_actual=sum(n(r["Cost of goods sold"]) for r in rows)
net_with_cost=sum(n(r["Net sales with cost recorded"]) for r in rows)
net_missing=sum(n(r["Net sales"]) for r in missing)
gross_missing=sum(n(r["Gross sales"]) for r in missing)
items_missing=sum(n(r["Net items sold"]) for r in missing)
csv_net=sum(n(r["Net sales"]) for r in rows)

# ---------- styles ----------
BOLD=Font(bold=True); H1=Font(bold=True,size=14); H2=Font(bold=True,size=11,color="FFFFFF")
FILL_H=PatternFill("solid",fgColor="1F2937"); FILL_SUB=PatternFill("solid",fgColor="E5E7EB"); FILL_IN=PatternFill("solid",fgColor="FEF3C7"); FILL_KEY=PatternFill("solid",fgColor="D1FAE5"); FILL_WARN=PatternFill("solid",fgColor="FEE2E2")
GBP='£#,##0;[Red]-£#,##0'; GBP2='£#,##0.00;[Red]-£#,##0.00'; PCT='0.0%'
thin=Side(style="thin",color="9CA3AF"); TOP=Border(top=thin)
def hdr(ws,row,vals,fill=FILL_H,font=H2):
    for i,v in enumerate(vals,1):
        c=ws.cell(row=row,column=i,value=v); c.fill=fill; c.font=font; c.alignment=Alignment(horizontal="center" if i>1 else "left",vertical="center",wrap_text=True)
def widths(ws,w):
    for i,x in enumerate(w,1): ws.column_dimensions[get_column_letter(i)].width=x

wb=Workbook()

# ================= Inputs & Assumptions =================
wsI=wb.active; wsI.title="Inputs & Assumptions"
wsI["A1"]="The Sword Stall – P&L inputs (Apr–Jun 2026)"; wsI["A1"].font=H1
wsI["A2"]="Yellow cells are editable assumptions. Everything on the P&L tabs recalculates from here. Status: CONFIRMED = from a hard source; CALL = from a client call (verify); ASSUMED = industry default; UNKNOWN = needs James."
wsI["A2"].alignment=Alignment(wrap_text=True); wsI.merge_cells("A2:F2"); wsI.row_dimensions[2].height=32
hdr(wsI,4,["Cost of delivery assumptions","Value","Unit","Status","Source / note",""])
inp=cfg["inputs"]
r=5; INPUT_CELLS={}
for it in inp:
    wsI.cell(row=r,column=1,value=it["name"]); c=wsI.cell(row=r,column=2,value=it["value"]); c.fill=FILL_IN
    c.number_format = PCT if it["unit"]=="% " else (GBP2 if it["unit"].startswith("£") else "0.00")
    wsI.cell(row=r,column=3,value=it["unit"]); wsI.cell(row=r,column=4,value=it["status"]); wsI.cell(row=r,column=5,value=it["note"])
    INPUT_CELLS[it["key"]]=f"'Inputs & Assumptions'!$B${r}"; r+=1
r+=1
hdr(wsI,r,["Operating expenses (monthly, ex VAT)","Apr 2026","May 2026","Jun 2026","Status","Source / note"]); r+=1
OPEX_START=r
for it in cfg["opex"]:
    wsI.cell(row=r,column=1,value=it["name"])
    for j,m in enumerate(MONTHS):
        v=it["monthly"][j] if isinstance(it["monthly"],list) else it["monthly"]
        c=wsI.cell(row=r,column=2+j,value=v); c.fill=FILL_IN; c.number_format=GBP
    wsI.cell(row=r,column=5,value=it["status"]); wsI.cell(row=r,column=6,value=it["note"])
    if it["status"]=="UNKNOWN":
        for j in range(1,5): wsI.cell(row=r,column=j).fill=FILL_WARN
    r+=1
OPEX_END=r-1
wsI.cell(row=r,column=1,value="Total operating expenses").font=BOLD
DIRECT_START=None
for j in range(3):
    col=get_column_letter(2+j); c=wsI.cell(row=r,column=2+j,value=f"=SUM({col}{OPEX_START}:{col}{OPEX_END})"); c.font=BOLD; c.number_format=GBP; c.border=TOP
OPEX_TOTAL_ROW=r
r+=2
hdr(wsI,r,["Other direct marketing (monthly, ex VAT) – sits with ad spend, scales with revenue","Apr 2026","May 2026","Jun 2026","Status","Source / note"]); r+=1
DIRECT_START=r
for it in cfg["direct"]:
    wsI.cell(row=r,column=1,value=it["name"])
    for j,m in enumerate(MONTHS):
        v=it["monthly"][j] if isinstance(it["monthly"],list) else it["monthly"]
        c=wsI.cell(row=r,column=2+j,value=v); c.fill=FILL_IN; c.number_format=GBP
    wsI.cell(row=r,column=5,value=it["status"]); wsI.cell(row=r,column=6,value=it["note"])
    if it["status"]=="UNKNOWN":
        for j in range(1,5): wsI.cell(row=r,column=j).fill=FILL_WARN
    r+=1
DIRECT_END=r-1
for c in wsI[f"A{DIRECT_START}:F{DIRECT_END}"]:
    for cell in c: cell.alignment=Alignment(wrap_text=True,vertical="top")
for c in wsI[f"A{OPEX_START}:F{OPEX_END}"]:
    for cell in c: cell.alignment=Alignment(wrap_text=True,vertical="top")
widths(wsI,[44,14,14,14,14,90])

# ================= Revenue tab =================
wsR=wb.create_sheet("Revenue (Shopify)")
wsR["A1"]="Shopify sales – Apr–Jun 2026 (from Shopify Analytics 'Total sales over time', daily, summed by month)"; wsR["A1"].font=H1
cols=["Orders","Gross sales","Discounts","Sales reversals","Net sales","Shipping charges","Duties","Additional fees","Taxes","Total sales"]
hdr(wsR,3,["Metric"]+[MLABEL[m] for m in MONTHS]+["Q2 total"])
REV={}
for i,cname in enumerate(cols):
    rr=4+i; wsR.cell(row=rr,column=1,value=cname)
    for j,m in enumerate(MONTHS):
        c=wsR.cell(row=rr,column=2+j,value=round(shop[m][cname],2)); c.number_format = '#,##0' if cname=="Orders" else GBP
    c=wsR.cell(row=rr,column=5,value=f"=SUM(B{rr}:D{rr})"); c.number_format='#,##0' if cname=="Orders" else GBP; c.font=BOLD
    REV[cname]=rr
rr=4+len(cols)+1
wsR.cell(row=rr,column=1,value="Revenue for P&L = Net sales + Shipping charges (ex VAT)").font=BOLD
for j in range(4):
    col=get_column_letter(2+j); c=wsR.cell(row=rr,column=2+j,value=f"={col}{REV['Net sales']}+{col}{REV['Shipping charges']}"); c.number_format=GBP; c.font=BOLD; c.fill=FILL_KEY
REV["Revenue"]=rr
rr+=1; wsR.cell(row=rr,column=1,value="AOV (net sales ÷ orders)")
for j in range(4):
    col=get_column_letter(2+j); c=wsR.cell(row=rr,column=2+j,value=f"={col}{REV['Net sales']}/{col}{REV['Orders']}"); c.number_format=GBP2
rr+=1; wsR.cell(row=rr,column=1,value="Returns as % of gross sales")
for j in range(4):
    col=get_column_letter(2+j); c=wsR.cell(row=rr,column=2+j,value=f"=-{col}{REV['Sales reversals']}/{col}{REV['Gross sales']}"); c.number_format=PCT
rr+=2
notes=["Definitions (Shopify): Net sales = Gross − Discounts − Sales reversals (returns/refunds). Total sales = Net + Shipping + Taxes. The P&L uses Net sales + Shipping charges, excluding VAT, per the 'one revenue definition' rule.",
 f"Cross-check: the product-variant COGS export sums to net sales of £{csv_net:,.0f} for the same period vs £{sum(shop[m]['Net sales'] for m in MONTHS):,.0f} here (difference is non-product items such as gift cards/tips and report timing).",
 "Source file: data/sword-stall/uploads/shopify/Total sales over time - 2024-08-01 - 2026-08-16.csv"]
for t in notes:
    wsR.cell(row=rr,column=1,value=t).alignment=Alignment(wrap_text=True); wsR.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=5); wsR.row_dimensions[rr].height=30; rr+=1
widths(wsR,[52,16,16,16,16])

# ================= Ad spend tab =================
wsA=wb.create_sheet("Ad Spend")
wsA["A1"]="Direct advertising – Apr–Jun 2026"; wsA["A1"].font=H1
hdr(wsA,3,["Platform / metric"]+[MLABEL[m] for m in MONTHS]+["Q2 total","Source"])
ADS={}
def put(rr,label,vals,fmt,src="",bold=False):
    wsA.cell(row=rr,column=1,value=label).font=Font(bold=bold)
    for j,v in enumerate(vals):
        c=wsA.cell(row=rr,column=2+j,value=v); c.number_format=fmt; c.font=Font(bold=bold)
    c=wsA.cell(row=rr,column=5,value=f"=SUM(B{rr}:D{rr})"); c.number_format=fmt; c.font=BOLD
    wsA.cell(row=rr,column=6,value=src)
put(4,"Google Ads spend (ex VAT)",[round(gads[m]['spend'],2) for m in MONTHS],GBP,f"Google Ads API → vendo.db gads_campaign_spend, account 2310522325 (last synced {max(g['synced'] for g in gads.values())[:10]})",True); ADS['google']=4
put(5,"  Google conversions",[round(gads[m]['conv'],0) for m in MONTHS],'#,##0',"Google Ads API")
put(6,"  Google conv. value (platform-attributed)",[round(gads[m]['value'],2) for m in MONTHS],GBP,"Google Ads API")
put(7,"Meta Ads spend (ex VAT)",[round(meta_tw[m],2) for m in MONTHS],GBP,"Triple Whale daily ad-spend feed (data/sword-stall/uploads/sword-stall.myshopify.com-all-2026-01-01-2026-08-15.csv). Cross-checked against Ads Manager campaign export to 16 Aug 2026 (row below). Meta API token expired 3 Sep – re-auth to pull live.",True); ADS['meta']=7
put(8,"  Meta spend per Ads Manager export (check)",[meta_am[m] for m in MONTHS],GBP,"Ads Manager campaign export, from outputs/analyses/2026-08-16-sword-stall-yoy.md")
put(9,"Total direct advertising",[f"={get_column_letter(2+j)}4+{get_column_letter(2+j)}7" for j in range(3)],GBP,"",True); ADS['total']=9
for j in range(4): wsA.cell(row=9,column=2+j).fill=FILL_KEY
wsA.cell(row=11,column=1,value="Note: per the P&L framework, 'direct advertising' = in-platform spend (+ influencer fees if any). Agency fees, creative production and software sit in operating expenses so that only this block scales with revenue.").alignment=Alignment(wrap_text=True)
wsA.merge_cells("A11:F11"); wsA.row_dimensions[11].height=32
wsA.cell(row=12,column=1,value="Ad spend is paid by the client directly to Google/Meta; figures exclude VAT (UK reverse charge). No influencer spend identified for the period – add a row if there was any.").alignment=Alignment(wrap_text=True)
wsA.merge_cells("A12:F12"); wsA.row_dimensions[12].height=32
widths(wsA,[42,14,14,14,14,110])

# ================= COGS actual tab =================
wsC=wb.create_sheet("COGS – Actual (Shopify)")
wsC["A1"]="Shopify 'Total sales by product variant with COGS' – 1 Apr to 30 Jun 2026 – as exported, no adjustments"; wsC["A1"].font=H1
fields=list(rows[0].keys())
hdr(wsC,3,fields+["Cost recorded?"])
for i,rw in enumerate(rows):
    rr=4+i
    for j,f in enumerate(fields):
        v=rw[f]
        if f not in ("Product title","Product variant title","Product variant SKU"): v=n(v)
        c=wsC.cell(row=rr,column=1+j,value=v)
        if f not in ("Product title","Product variant title","Product variant SKU","Net items sold"): c.number_format=GBP2
    flag="NO" if (n(rw["Cost of goods sold"])==0 and n(rw["Net items sold"])>0) else "yes"
    c=wsC.cell(row=rr,column=len(fields)+1,value=flag)
    if flag=="NO":
        for j in range(len(fields)+1): wsC.cell(row=rr,column=1+j).fill=FILL_WARN
last=4+len(rows)-1; tr=last+2
wsC.cell(row=tr,column=1,value="TOTAL").font=BOLD
for j,f in enumerate(fields):
    if f in ("Product title","Product variant title","Product variant SKU"): continue
    col=get_column_letter(1+j); c=wsC.cell(row=tr,column=1+j,value=f"=SUM({col}4:{col}{last})"); c.font=BOLD; c.number_format=GBP2 if f!="Net items sold" else '#,##0'; c.border=TOP
COGS_ACT={"total_row":tr,"cogs_col":get_column_letter(1+fields.index("Cost of goods sold")),"net_col":get_column_letter(1+fields.index("Net sales")),"nwc_col":get_column_letter(1+fields.index("Net sales with cost recorded"))}
wsC.cell(row=tr+2,column=1,value=f"{len(missing)} variants ({int(items_missing)} units, £{net_missing:,.0f} net sales) sold in the period with no cost recorded in Shopify – highlighted red. Their COGS is counted as £0 on this tab, which overstates margin. See 'COGS – Estimated' for the gap filled.").alignment=Alignment(wrap_text=True)
wsC.merge_cells(start_row=tr+2,start_column=1,end_row=tr+2,end_column=8); wsC.row_dimensions[tr+2].height=32
wsC.freeze_panes="A4"; widths(wsC,[60,18,18,10,13,12,12,13,12,13,14,16,12])

# ================= COGS estimated tab =================
wsE=wb.create_sheet("COGS – Estimated")
wsE["A1"]="Estimated cost for the variants with no COGS in Shopify"; wsE["A1"].font=H1
wsE["A2"]=("Method: (1) memberships/gift cards/digital = £0 cost. (1b) Statues/busts/Hot Toys/Weta/life-size props: James said on the 14 May 2026 call that statues carry ~10% markup vs ~70% on swords, so cost = price ÷ 1.10. (2) If the SKU prefix (supplier code, e.g. VX, UC, HK, BS) has ≥5 variants with a recorded cost, apply that supplier's average COGS-to-gross-sales ratio. "
           f"(3) Otherwise apply the store-wide ratio of {overall_ratio:.1%} (recorded COGS ÷ gross sales across all {sum(v[2] for v in ref.values())} variants with cost). Ratios are editable in column H; the estimate recalculates.")
wsE["A2"].alignment=Alignment(wrap_text=True); wsE.merge_cells("A2:J2"); wsE.row_dimensions[2].height=48
hdr(wsE,4,["Product title","Variant","SKU","Units sold","Gross sales","Net sales","Method","Ratio (COGS ÷ gross)","Estimated COGS","Confidence"])
for i,rw in enumerate(sorted(missing,key=lambda r:-n(r["Gross sales"]))):
    rr=5+i; est,method,ratio=estimate(rw)
    conf="High" if ratio==0 else ("Medium" if ("prefix" in method or "Collectible" in method) else "Low")
    vals=[rw["Product title"],rw["Product variant title"],rw["Product variant SKU"],n(rw["Net items sold"]),n(rw["Gross sales"]),n(rw["Net sales"]),method,round(ratio,4),f"=E{rr}*H{rr}",conf]
    for j,v in enumerate(vals):
        c=wsE.cell(row=rr,column=1+j,value=v)
        if j in (4,5,8): c.number_format=GBP2
        if j==7: c.number_format=PCT; c.fill=FILL_IN
lastE=5+len(missing)-1; te=lastE+2
wsE.cell(row=te,column=1,value="TOTAL estimated COGS for missing variants").font=BOLD
for col in "DEFI":
    c=wsE.cell(row=te,column=ord(col)-64,value=f"=SUM({col}5:{col}{lastE})"); c.font=BOLD; c.border=TOP; c.number_format=GBP2 if col!="D" else '#,##0'
COGS_EST_TOTAL=f"'COGS – Estimated'!$I${te}"
te+=2
hdr(wsE,te,["Supplier prefix","Variants with cost","Gross sales (with cost)","Recorded COGS","Ratio","Used for missing rows (count)","","","",""],fill=FILL_SUB,font=BOLD); te+=1
miss_by=defaultdict(int)
for rw in missing: miss_by[pref(rw["Product variant SKU"])]+=1
for p,v in sorted(ref.items(),key=lambda kv:-kv[1][0]):
    wsE.cell(row=te,column=1,value=p); wsE.cell(row=te,column=2,value=v[2]); wsE.cell(row=te,column=3,value=round(v[0],2)).number_format=GBP; wsE.cell(row=te,column=4,value=round(v[1],2)).number_format=GBP; wsE.cell(row=te,column=5,value=(round(v[1]/v[0],4) if v[0] else 0)).number_format=PCT; wsE.cell(row=te,column=6,value=miss_by.get(p,0) if v[2]>=MIN_ROWS else f"{miss_by.get(p,0)} (too few refs → store avg)")
    te+=1
wsE.freeze_panes="A5"; widths(wsE,[60,16,18,10,13,13,52,16,15,11])

# ================= P&L tabs =================
def build_pnl(name,use_estimate):
    ws=wb.create_sheet(name)
    ws["A1"]=f"The Sword Stall – P&L, Apr–Jun 2026 ({'recorded COGS only – cost gaps counted as £0' if not use_estimate else 'recorded COGS + estimate for the variants missing a cost'})"; ws["A1"].font=H1
    ws["A2"]="Structure: Net sales (+ shipping collected) → Cost of delivery → GROSS MARGIN → Direct advertising → CONTRIBUTION MARGIN (CM2) → Operating expenses → NET PROFIT. All ex VAT. Monthly COGS is the Q2 total allocated pro-rata to each month's net sales (Shopify's COGS export is a quarter total); the Q2 column is the actual figure."
    ws["A2"].alignment=Alignment(wrap_text=True); ws.merge_cells("A2:J2"); ws.row_dimensions[2].height=44
    hdr(ws,4,["",MLABEL["2026-04"],"% rev",MLABEL["2026-05"],"% rev",MLABEL["2026-06"],"% rev","Q2 2026","% rev","Basis / source"])
    R="'Revenue (Shopify)'!"; A="'Ad Spend'!"; I="'Inputs & Assumptions'!"
    mcols=["B","C","D","E"]  # month cols on Revenue/Ad tabs (B,C,D + E total)
    pcols=["B","D","F","H"]; pct=["C","E","G","I"]
    rr=5; ROW={}
    def line(label,formulas,fmt=GBP,bold=False,fill=None,note="",pct_of=None,sign=1):
        nonlocal rr
        ws.cell(row=rr,column=1,value=label).font=Font(bold=bold)
        for j in range(4):
            c=ws[f"{pcols[j]}{rr}"]; c.value=formulas[j]; c.number_format=fmt; c.font=Font(bold=bold)
            if fill: c.fill=fill
            if pct_of:
                p=ws[f"{pct[j]}{rr}"]; p.value=f"=IFERROR({pcols[j]}{rr}/{pcols[j]}{pct_of},0)"; p.number_format=PCT; p.font=Font(bold=bold,color="6B7280")
                if fill: p.fill=fill
        ws.cell(row=rr,column=10,value=note).alignment=Alignment(wrap_text=True,vertical="top")
        ROW[label]=rr; rr+=1; return rr-1
    def section(t):
        nonlocal rr
        c=ws.cell(row=rr,column=1,value=t); c.font=BOLD; c.fill=FILL_SUB
        for j in range(2,11): ws.cell(row=rr,column=j).fill=FILL_SUB
        rr+=1
    section("1. REVENUE")
    ns=line("Net sales (after discounts & returns)",[f"={R}{mcols[j]}{REV['Net sales']}" for j in range(4)],note="Shopify Analytics")
    sh=line("Shipping collected at checkout",[f"={R}{mcols[j]}{REV['Shipping charges']}" for j in range(4)],note="Shopify Analytics")
    rev=line("TOTAL REVENUE (ex VAT)",[f"={pcols[j]}{ns}+{pcols[j]}{sh}" for j in range(4)],bold=True,fill=FILL_KEY)
    for j in range(4): ws[f"{pct[j]}{rev}"].value=1; ws[f"{pct[j]}{rev}"].number_format=PCT
    orders=line("Orders",[f"={R}{mcols[j]}{REV['Orders']}" for j in range(4)],fmt='#,##0',note="Shopify Analytics")
    rr+=1; section("2. COST OF DELIVERY")
    CA=COGS_ACT
    q2_cogs=f"'COGS – Actual (Shopify)'!{CA['cogs_col']}{CA['total_row']}" + (f"+{COGS_EST_TOTAL}" if use_estimate else "")
    cogs=line("Cost of goods sold",[f"=({q2_cogs})*{pcols[j]}{ns}/$H${ns}" if j<3 else f"={q2_cogs}" for j in range(4)],pct_of=rev,
              note=("Shopify recorded COGS (£{:,.0f})".format(cogs_actual)+(" + estimate for variants with no cost (see 'COGS – Estimated')" if use_estimate else f" only. {len(missing)} variants / £{net_missing:,.0f} net sales have NO cost recorded and are counted at £0 here – margin is overstated."))+" Monthly = pro-rata to net sales.")
    tf=line("Payment processing / transaction fees",[f"={I.replace(chr(39),chr(39))}$B$5*{R}{mcols[j]}{REV['Total sales']}" for j in range(4)],pct_of=rev,note="Fee % (Inputs) × Shopify Total sales (the amount actually charged incl. VAT & shipping). Refund fees are usually not returned by processors – not modelled.")
    # fix tf formula to use the named input cell
    for j in range(4): ws[f"{pcols[j]}{tf}"].value=f"={INPUT_CELLS['txn_fee']}*{R}{mcols[j]}{REV['Total sales']}"
    shp=line("Shipping & fulfilment (carrier cost)",[f"={INPUT_CELLS['ship_per_order']}*{pcols[j]}{orders}" for j in range(4)],pct_of=rev,note="Cost per order (Inputs) × orders. Should be the actual carrier invoice ÷ orders shipped – get the DPD/DX invoices.")
    pkg=line("Packaging & consumables",[f"={INPUT_CELLS['pack_per_order']}*{pcols[j]}{orders}" for j in range(4)],pct_of=rev,note="Per-order packaging (Inputs). UNKNOWN – set to 0 until James confirms.")
    cod=line("TOTAL COST OF DELIVERY",[f"=SUM({pcols[j]}{cogs}:{pcols[j]}{pkg})" for j in range(4)],bold=True,pct_of=rev)
    rr+=1
    gm=line("GROSS MARGIN (CM1)",[f"={pcols[j]}{rev}-{pcols[j]}{cod}" for j in range(4)],bold=True,fill=FILL_KEY,pct_of=rev)
    rr+=1; section("3. DIRECT ADVERTISING & MARKETING")
    g=line("Google Ads",[f"={A}{mcols[j]}{ADS['google']}" for j in range(4)],pct_of=rev,note="Google Ads API")
    mt=line("Meta Ads",[f"={A}{mcols[j]}{ADS['meta']}" for j in range(4)],pct_of=rev,note="Triple Whale feed, cross-checked vs Ads Manager export")
    drows=[]
    for k,it in enumerate(cfg["direct"]):
        irow=DIRECT_START+k
        d=line(it["name"],[f"={I}{'BCD'[j]}{irow}" for j in range(3)]+[f"=B{rr}+D{rr}+F{rr}"],pct_of=rev,note=f"{it['status']} – {it['note'][:160]}")
        drows.append(d)
    adv=line("TOTAL DIRECT ADVERTISING",[f"=SUM({pcols[j]}{g}:{pcols[j]}{drows[-1]})" for j in range(4)],bold=True,pct_of=rev)
    mer=line("MER (revenue ÷ ad spend)",[f"=IFERROR({pcols[j]}{rev}/{pcols[j]}{adv},0)" for j in range(4)],fmt='0.00"x"',note="Marketing efficiency ratio on TOTAL revenue, not platform-attributed revenue.")
    rr+=1
    cm=line("CONTRIBUTION MARGIN (CM2)",[f"={pcols[j]}{gm}-{pcols[j]}{adv}" for j in range(4)],bold=True,fill=FILL_KEY,pct_of=rev)
    rr+=1; section("4. OPERATING EXPENSES (from Inputs tab)")
    opex_rows=[]
    for k,it in enumerate(cfg["opex"]):
        irow=OPEX_START+k
        o=line(it["name"],[f"={I}{'BCD'[j]}{irow}" for j in range(3)]+[f"=B{rr}+D{rr}+F{rr}"],pct_of=rev,note=f"{it['status']} – {it['note'][:140]}")
        opex_rows.append(o)
    opx=line("TOTAL OPERATING EXPENSES",[f"=SUM({pcols[j]}{opex_rows[0]}:{pcols[j]}{opex_rows[-1]})" for j in range(4)],bold=True,pct_of=rev)
    rr+=1
    npf=line("NET PROFIT (CM3 / EBITDA)",[f"={pcols[j]}{cm}-{pcols[j]}{opx}" for j in range(4)],bold=True,fill=FILL_KEY,pct_of=rev)
    rr+=2; section("MEMO LINES")
    line("Marketing all-in (ads + Vendo fee) % of revenue",[f"=IFERROR(({pcols[j]}{adv}+{pcols[j]}{ROW.get('Vendo agency fee (Digital Marketing Suite)',adv)})/{pcols[j]}{rev},0)" for j in range(4)],fmt=PCT,note="What James is likely calling 'marketing'. Compare with his 27% claim.")
    line("Platform-attributed ROAS – Google",[f"=IFERROR({A}{mcols[j]}6/{A}{mcols[j]}4,0)" for j in range(4)],fmt='0.00"x"',note="Google conversion value ÷ spend. Platform-attributed, for reference only.")
    line("Blended net margin per order",[f"=IFERROR({pcols[j]}{npf}/{pcols[j]}{orders},0)" for j in range(4)],fmt=GBP2)
    line("Returns (sales reversals) – already netted in revenue",[f"=-{R}{mcols[j]}{REV['Sales reversals']}" for j in range(4)],pct_of=rev,note="Shown for visibility: refunds are deducted before Net sales. 8–11% of gross is the profitability lever outside ads.")
    for j in range(4): ws[f"{pct[j]}{ROW['Returns (sales reversals) – already netted in revenue']}"].value=f"=IFERROR(-{R}{mcols[j]}{REV['Sales reversals']}/{R}{mcols[j]}{REV['Gross sales']},0)"
    ws.freeze_panes="B5"; widths(ws,[46,13,8,13,8,13,8,14,8,80])
    for row in ws.iter_rows(min_row=5,max_row=rr,min_col=10,max_col=10):
        for c in row: c.font=Font(size=9,color="374151")
    return ws
build_pnl("P&L – Actual COGS",False)
build_pnl("P&L – Estimated COGS",True)

# ================= Open questions =================
wsQ=wb.create_sheet("Open questions for James")
wsQ["A1"]="Things I could not source – needed to finish the P&L"; wsQ["A1"].font=H1
hdr(wsQ,3,["#","Question","Why it matters","Current placeholder"])
for i,q in enumerate(cfg["questions"]):
    for j,v in enumerate([i+1]+q): c=wsQ.cell(row=4+i,column=1+j,value=v); c.alignment=Alignment(wrap_text=True,vertical="top")
widths(wsQ,[4,70,60,40])

# order sheets
order=["P&L – Estimated COGS","P&L – Actual COGS","Inputs & Assumptions","Revenue (Shopify)","Ad Spend","COGS – Actual (Shopify)","COGS – Estimated","Open questions for James"]
wb._sheets=[wb[s] for s in order]
wb.save(OUT)
print("saved",OUT)
print(f"COGS recorded £{cogs_actual:,.0f}; estimated gap £{est_total:,.0f}; net w/ cost £{net_with_cost:,.0f}; missing net £{net_missing:,.0f}; overall ratio {overall_ratio:.3%}")
print("Google",{m:round(gads[m]['spend']) for m in MONTHS},"Meta TW",{m:round(meta_tw[m]) for m in MONTHS})
